from datetime import datetime
import io

import pandas as pd
from io import BytesIO, StringIO
from flask import Response
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def safe_str(value):
    """Convert any value to string safely"""
    if value is None:
        return ''
    if isinstance(value, ObjectId):
        return str(value)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    return str(value)


def export_to_excel(registrations):
    """Export registrations to Excel format"""
    # Convert to DataFrame
    data = []
    for reg in registrations:
        data.append(
            {
                "ID": str(reg.get("_id", "")),
                "Full Name": reg.get("full_name", ""),
                "Email": reg.get("email", ""),
                "Institution": reg.get("institution", ""),
                "Segment": reg.get("segment_name", ""),
                "Category": reg.get("category", ""),
                "Submission Link": reg.get("submission_link", ""),
                "CA Ref": reg.get("ca_ref", ""),
                "Bkash Number": reg.get("bkash_number", ""),
                "Transaction ID": reg.get("transaction_id", ""),
                "Receipt": reg.get("receipt", ""),
                "Verified": "Yes" if reg.get("verified") else "No",
                "Registration Date": reg.get("registration_date", ""),
                "Verified At": reg.get("verified_at", ""),
                "Firebase uid": reg.get("firebase_uid", ""),
                "IP Address": reg.get("ip_address", ""),
            }
        )

    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = BytesIO()
    
    # Use openpyxl directly without pandas ExcelWriter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=registrations.xlsx'}
    )


def export_to_csv(registrations):
    """Export registrations to CSV format with proper formatting"""
    # Convert to DataFrame with safe string conversion
    data = []
    for reg in registrations:
        # Format registration date
        reg_date = reg.get('registration_date')
        if isinstance(reg_date, datetime):
            formatted_reg_date = reg_date.strftime('%Y-%m-%d %H:%M:%S')
        else:
            formatted_reg_date = safe_str(reg_date)
        
        # Format verified date
        verified_at = reg.get('verified_at')
        if isinstance(verified_at, datetime):
            formatted_verified_at = verified_at.strftime('%Y-%m-%d %H:%M:%S')
        else:
            formatted_verified_at = safe_str(verified_at)
        
        data.append({
            "ID": safe_str(reg.get("_id")),
            "Full Name": safe_str(reg.get("full_name")),
            "Email": safe_str(reg.get("email")),
            "Institution": safe_str(reg.get("institution")),
            "Segment": safe_str(reg.get("segment_name")),
            "Category": safe_str(reg.get("category")),
            "Submission Link": safe_str(reg.get("submission_link")),
            "CA Ref": safe_str(reg.get("ca_ref")),
            "Bkash Number": safe_str(reg.get("bkash_number")),
            "Transaction ID": safe_str(reg.get("transaction_id")),
            "Receipt": safe_str(reg.get("receipt")),
            "Verified": "Yes" if reg.get("verified") else "No",
            "Registration Date": formatted_reg_date,
            "Verified At": formatted_verified_at,
            "Firebase uid": safe_str(reg.get("firebase_uid")),
            "IP Address": safe_str(reg.get("ip_address")),
        })

    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Handle empty DataFrame
    if df.empty:
        df = pd.DataFrame(columns=[
            "ID", "Full Name", "Email", "Institution", "Segment",
            "Category", "Submission Link", "CA Ref", "Bkash Number",
            "Transaction ID", "Verified", "Registration Date",
            "Verified At", "Firebase uid", "IP Address"
        ])
    
    # Create CSV with UTF-8 encoding
    output = io.StringIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    
    # Convert to bytes for response with BOM for Excel
    csv_bytes = output.getvalue().encode('utf-8-sig')
    
    return Response(
        csv_bytes,
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=registrations.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )


def export_ca_to_csv(ca_data):
    """Export CA registrations to CSV format with proper formatting"""
    # Convert to DataFrame with safe string conversion
    data = []
    for ca in ca_data:
        # Format registration date
        reg_date = ca.get('registration_date')
        if isinstance(reg_date, datetime):
            formatted_reg_date = reg_date.strftime('%Y-%m-%d %H:%M:%S')
        else:
            formatted_reg_date = safe_str(reg_date)
        
        data.append({
            "ID": safe_str(ca.get("_id")),
            "CA Code": safe_str(ca.get("ca_code")),
            "Profile": safe_str(ca.get("profile_picture")),
            "Full Name": safe_str(ca.get("full_name")),
            "Email": safe_str(ca.get("email")),
            "Institution": safe_str(ca.get("institution")),
            "Class": safe_str(ca.get("class")),
            "Phone Number": safe_str(ca.get("phone")),
            "Status": safe_str(ca.get("status")),
            "Facebook": safe_str(ca.get("facebook_link")),
            "Why CA": safe_str(ca.get("why_ca")),
            "Registration Date": formatted_reg_date,
            "Firebase uid": safe_str(ca.get("firebase_uid")),
            "IP Address": safe_str(ca.get("ip_address")),
        })

    # Convert to DataFrame
    df = pd.DataFrame(data)
    
    # Handle empty DataFrame
    if df.empty:
        df = pd.DataFrame(columns=[
            "ID", "CA Code", "Profile", "Full Name", "Email", 
            "Institution", "Class", "Phone Number", "Status", 
            "Why CA", "Registration Date", "Firebase uid", "IP Address"
        ])
    
    # Create CSV with UTF-8 encoding
    output = io.StringIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    
    # Convert to bytes for response with BOM for Excel
    csv_bytes = output.getvalue().encode('utf-8-sig')
    
    return Response(
        csv_bytes,
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=ca_registrations.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )


def export_ca_to_excel(ca_data):
    """Export registrations to CSV format"""
    # Convert to DataFrame
    data = []
    for ca in ca_data:
        data.append(
            {
                "ID": str(ca.get("_id", "")),
                "CA Code": ca.get("ca_code", ""),
                "Profile": ca.get("profile_picture", ""),
                "Full Name": ca.get("full_name", ""),
                "Email": ca.get("email", ""),
                "Institution": ca.get("institution", ""),
                "Class": ca.get("class", ""),
                "Phone Number": ca.get("phone", ""),
                "Status": ca.get("status"),
                "Facebook": ca.get("facebook_link"),
                "Why CA": ca.get("why_ca", ""),
                "Registration Date": ca.get("registration_date", ""),
                "Firebase uid": ca.get("firebase_uid", ""),
                "IP Address": ca.get("ip_address", ""),
            }
        )

    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = BytesIO()
    
    # Use openpyxl directly without pandas ExcelWriter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CARegistrations"
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=caregistrations.xlsx'}
    )


def export_bob_to_excel(bob_data):
    """Export BoB registrations to Excel format matching CA export style"""
    # Convert to DataFrame
    data = []
    for bob in bob_data:
        # Format members as a readable string
        members_str = ""
        for member in bob.get('members', []):
            members_str += f"{member.get('position', '')}. {member.get('name', '')} - {member.get('role', '')}\n"
        members_str = members_str.strip()
        
        data.append({
            "ID": str(bob.get("_id", "")),
            "Band Name": bob.get("band_name", ""),
            "Email": bob.get("email", ""),
            "Institution": bob.get("institution", ""),
            "Genre": bob.get("band_genre", ""),
            "Member Count": bob.get("member_count", 0),
            "Members": members_str,
            "Jamming Clip": bob.get("jamming_clip", ""),
            "CA Reference": bob.get("ca_reference", "") if bob.get("ca_reference") else "",
            "Verified": "Yes" if bob.get("verified") else "No",
            "Status": bob.get("status", "pending"),
            "Registration Date": bob.get("registration_date", ""),
            "Firebase uid": bob.get("firebase_uid", ""),
            "User ID": str(bob.get("user_id", "")),
            "IP Address": bob.get("ip_address", ""),
        })

    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    # Create workbook directly with openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "BoBRegistrations"
    
    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)
    
    # Write data
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Cap width at 50 and add padding
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Enable text wrapping for Members column (column G)
    if 'G' in ws.column_dimensions:
        ws.column_dimensions['G'].width = 50  # Make Members column wider
    
    wb.save(output)
    output.seek(0)
    
    # Create filename with timestamp
    from datetime import datetime
    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"bob_registrations_{timestamp}.xlsx"
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename}',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8'
        }
    )